"""Generate code/factory-control-codes.xlsx — an editable list of the
Factory-Control (exception) code rule used in compare.py.

Factory-Control 판정 로직(compare.py `_is_fc`)은 두 부분으로 구성됩니다:
    1) 접두어 규칙 — 코드 첫 글자가 I / O / Z / U 이면 자동으로 Factory Control
    2) 개별 추가 코드 — 위 접두어에 해당하지 않지만 명시적으로 포함하는 코드
       (DEFAULT_EXCEPT_CODES 의 추가분: DUP0, A0B, E0D, E0Q, J7G)

Layout (편집/추가 가능하도록):
    A = 유형(Type)      접두어(prefix) / 개별코드(code)
    B = 코드·접두어      e.g. I  /  DUP0
    C = 설명(Description)
    D = 비고(Note)

Source: backend/compare.py (DEFAULT_EXCEPT_CODES) + option_codes.OPTION_CODE_MAP.
Run:  python backend/build_factory_control_xlsx.py
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
sys.path.insert(0, str(HERE))

from option_codes import OPTION_CODE_MAP  # noqa: E402

OUT = ROOT / 'code' / 'factory-control-codes.xlsx'

# Mirror of compare.py: I/O/Z/U prefixes + these explicit extras.
FC_PREFIXES = ['I', 'O', 'Z', 'U']
FC_EXTRA_CODES = ['DUP0', 'A0B', 'E0D', 'E0Q', 'J7G']

NAVY = '1F4E79'
PREFIX_FILL = 'DDEBF7'   # blue-ish
CODE_FILL = 'FCE4D6'     # orange-ish

HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
BOLD = Font(bold=True)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center')


def _header(ws, headers):
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER


def build():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- Sheet 1: the editable rule (prefixes + extra codes) ------------------
    ws = wb.active
    ws.title = 'FactoryControl'
    _header(ws, ['유형(Type)', '코드·접두어(Code/Prefix)', '설명(Description)', '비고(Note)'])

    r = 2
    for p in FC_PREFIXES:
        ws.cell(row=r, column=1, value='접두어(prefix)')
        ws.cell(row=r, column=2, value=p).font = BOLD
        ws.cell(row=r, column=3, value=f"'{p}' 로 시작하는 모든 코드가 Factory Control")
        ws.cell(row=r, column=4, value='접두어 규칙')
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if col == 1:
                cell.fill = PatternFill('solid', fgColor=PREFIX_FILL)
        r += 1

    for code in FC_EXTRA_CODES:
        ws.cell(row=r, column=1, value='개별코드(code)')
        ws.cell(row=r, column=2, value=code).font = BOLD
        ws.cell(row=r, column=3, value=OPTION_CODE_MAP.get(code, ''))
        ws.cell(row=r, column=4, value='추가 지정 코드')
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if col == 1:
                cell.fill = PatternFill('solid', fgColor=CODE_FILL)
        r += 1

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 56
    ws.column_dimensions['D'].width = 20
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:D{r - 1}'

    # Note under the table.
    ws.cell(row=r + 1, column=1,
            value='※ 접두어 규칙(I/O/Z/U)에 걸리는 코드는 아래 "접두어_해당코드" 시트에서 확인. '
                  '개별코드 행을 추가/삭제하면 예외 목록이 바뀝니다.').font = Font(italic=True, color='777777')

    # --- Sheet 2: all codes matched by the I/O/Z/U prefix rule (reference) -----
    ref = wb.create_sheet('접두어_해당코드(Ref)')
    _header(ref, ['접두어(Prefix)', '코드(Code)', '설명(Description)'])
    matched = sorted(
        (c for c in OPTION_CODE_MAP if c and c[0] in FC_PREFIXES),
        key=lambda c: (c[0], c),
    )
    rr = 2
    for c in matched:
        ref.cell(row=rr, column=1, value=c[0])
        ref.cell(row=rr, column=2, value=c).font = BOLD
        ref.cell(row=rr, column=3, value=OPTION_CODE_MAP.get(c, ''))
        for col in range(1, 4):
            cell = ref.cell(row=rr, column=col)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
        rr += 1
    ref.column_dimensions['A'].width = 12
    ref.column_dimensions['B'].width = 12
    ref.column_dimensions['C'].width = 64
    ref.freeze_panes = 'A2'
    ref.auto_filter.ref = f'A1:C{rr - 1}'

    wb.save(OUT)
    print(f'Wrote {OUT}  ({len(FC_PREFIXES)} prefixes + {len(FC_EXTRA_CODES)} extra codes; '
          f'{len(matched)} prefix-matched codes for reference)')


if __name__ == '__main__':
    build()
