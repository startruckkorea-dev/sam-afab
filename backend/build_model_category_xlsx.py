"""Generate code/model-category.xlsx — Baumuster prefix -> tractor/rigid/tipper.

Rows are discovered from docs/data.json (every 6-digit Baumuster prefix seen, with
its example models / vehicle / subcategories). The category column is pre-filled
with the built-in rule (model_category.classify_prefix), but ANY category already
set in an existing workbook is preserved — so running this adds newly-seen prefixes
without clobbering your edits.

Read back at build time by model_category.load_model_category to decide which
mandatory codes apply per vehicle.

Run:  python backend/build_model_category_xlsx.py   (or refreshed by build_data.py)
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
sys.path.insert(0, str(HERE))

from model_category import classify_prefix, load_model_category  # noqa: E402

DATA_PATH = ROOT / 'docs' / 'data.json'
OUT = ROOT / 'code' / 'model-category.xlsx'

NAVY = '1F4E79'
CAT_FILL = {'tractor': 'DDEBF7', 'rigid': 'E2EFDA', 'tipper': 'FFF2CC'}

HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
BOLD = Font(bold=True)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center')


def _prefix_rows():
    """{prefix: {models,veh,subs}} from docs/data.json."""
    pref = defaultdict(lambda: {'models': set(), 'veh': set(), 'subs': set()})
    if not DATA_PATH.exists():
        return pref
    data = json.loads(DATA_PATH.read_text(encoding='utf-8'))
    for r in data.get('rows', []):
        m = re.match(r'(\d{6})', str(r.get('Baumuster', '') or '').strip())
        if not m:
            continue
        p = pref[m.group(1)]
        for key, col in (('models', 'Model(WINGS)'), ('veh', 'Vehicle'), ('subs', 'Subcategory (ID)')):
            v = str(r.get(col, '') or '').strip()
            if v:
                p[key].add(v)
    return pref


def build():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    existing = load_model_category()          # preserve prior category edits
    pref = _prefix_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = 'ModelCategory'
    headers = ['Baumuster접두어(6자리)', '카테고리(Category)', '예시모델(Example)',
               '차종(Vehicle)', 'Subcategory', '비고(Note)']
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER

    r = 2
    for prefix in sorted(pref):
        d = pref[prefix]
        cat = existing.get(prefix) or classify_prefix(prefix)
        ws.cell(row=r, column=1, value=prefix).font = BOLD
        ws.cell(row=r, column=2, value=cat)
        ws.cell(row=r, column=3, value=', '.join(sorted(d['models'])))
        ws.cell(row=r, column=4, value=', '.join(sorted(d['veh'])))
        ws.cell(row=r, column=5, value=', '.join(sorted(d['subs'])))
        ws.cell(row=r, column=6, value='')
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
        if cat in CAT_FILL:
            ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=CAT_FILL[cat])
        r += 1

    for col, w in zip('ABCDEF', [20, 16, 34, 12, 14, 30]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:F{max(r - 1, 1)}'

    # Legend
    lg = wb.create_sheet('설명(Legend)')
    lg.column_dimensions['A'].width = 16
    lg.column_dimensions['B'].width = 74
    lg.append(['항목', '설명'])
    for i in (1, 2):
        c = lg.cell(row=1, column=i)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
    for k, v in [
        ('tractor', '트랙터 (Sattelzugmaschine, 예: LS/S). BM 963425, 964416, 963403, 964424 …'),
        ('rigid', '리지드 (일반 화물, 예: L). 그 외 964XXX'),
        ('tipper', '티퍼/덤퍼 (Kipper, 예: K). BM 964230, 964214, 964231 …'),
        ('', ''),
        ('사용', 'B열 카테고리를 수정하면 다음 빌드부터 mandatory 판정에 반영. '
                '표에 없는 BM은 규칙(964xxx=rigid 등)으로 자동 분류.'),
    ]:
        lg.append([k, v])
    for row in lg.iter_rows(min_row=2, max_row=lg.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = WRAP_TOP
            cell.border = BORDER
    lg.cell(row=2, column=1).font = BOLD

    wb.save(OUT)
    print(f'Wrote {OUT}  ({r - 2} Baumuster prefixes)')


if __name__ == '__main__':
    build()
