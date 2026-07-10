"""Model-conversion rules — single source of truth, loaded from docs/rules.json.

The rules used to be hardcoded across sam_parser.py / compare.py. They now live
in docs/rules.json so they can be viewed/edited from the web UI (and later from a
SharePoint-hosted file). If the JSON is missing or a key is absent, the built-in
DEFAULTS below are used, so the build never breaks.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
RULES_PATH = ROOT / 'docs' / 'rules.json'
XLSX_PATH = ROOT / 'model_rules' / 'model_mapping.xlsx'

# model_mapping.xlsx editable sheets -> rules key. The '인식모델_대조표' sheet is an
# auto-generated recognition view and is ignored here. Matching is data-driven from
# the SAM files, so only the manual alias-override sheet feeds back into the rules.
_XLSX_MAP_SHEETS: dict = {}
_XLSX_LIST_SHEETS = {
    '매칭_별칭(수동)': 'reverse_aliases',
}

# Built-in defaults == the original hardcoded values (fallback only).
DEFAULTS = {
    'normalize_historic': {
        '3253': '4153', '4140': '4440', '2643': '3343',
        '2851': '2651', '2135': '1835', '2863': '2663', '2853': '2653',
    },
    'normalize_28xx_to_26xx': True,
    'reverse_aliases': {'3253': ['4153']},
    'previous_model': {
        '4453': '4153', '4153': '3253', '3343': '2643', '2853': '2663', '2851': '2661',
    },
    'current_model': {
        '4453': '4463', '4153': '4163', '3343': '3363', '2853': '2863', '2851': '2861',
    },
    'wings_display_replace': {
        '4140': '4440', '2651 LS': '2851 LS', '2653 LS': '2853 LS',
        '2663 LS': '2863 LS', '2643 A': '3343 A',
    },
    'vehicle_keywords': {
        'Actros-L': ['2651', '2851', '2653', '2853', '2663', '2863'],
        'Actros': ['3363'],
        'Arocs': ['2643', '3343', '4153', '4453', '3253', '2135', '4440', '4140'],
    },
    # Manual display overrides, keyed by WINGS model. { wings: {'baumuster','now'} }.
    # Wins over auto-recognized SAM Baumuster/now. Edited in model_mapping.xlsx '수동매핑'.
    'manual_map': {},
}


def load_rules_from_xlsx(path: Path | None = None) -> dict:
    """Parse model_rules/model_mapping.xlsx rule sheets into a rules dict.

    Returns only the keys present in the workbook (partial). openpyxl is imported
    lazily and every failure is swallowed so the build never breaks on a bad file.
    """
    p = path or XLSX_PATH
    out: dict = {}
    try:
        from openpyxl import load_workbook  # lazy: optional at build time
        wb = load_workbook(Path(p), read_only=True, data_only=True)
    except Exception:
        return out

    def _rows(ws):
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  # header
                continue
            if not row:
                continue
            k = '' if row[0] is None else str(row[0]).strip()
            v = '' if len(row) < 2 or row[1] is None else str(row[1]).strip()
            if k:
                yield k, v

    for sheet, key in _XLSX_MAP_SHEETS.items():
        if sheet in wb.sheetnames:
            out[key] = {k: v for k, v in _rows(wb[sheet])}
    for sheet, key in _XLSX_LIST_SHEETS.items():
        if sheet in wb.sheetnames:
            out[key] = {k: [s.strip() for s in v.split(',') if s.strip()]
                        for k, v in _rows(wb[sheet])}
    if '옵션' in wb.sheetnames:
        for k, v in _rows(wb['옵션']):
            if k == 'normalize_28xx_to_26xx':
                out['normalize_28xx_to_26xx'] = str(v).strip().lower() in ('true', '1', 'yes', 'y', 'on')

    # 수동매핑: WINGS 모델 | SAM Baumuster(원본) | SAM now(수정) | SAM 파일 지정(선택)
    if '수동매핑' in wb.sheetnames:
        mm = {}
        for i, row in enumerate(wb['수동매핑'].iter_rows(values_only=True)):
            if i == 0 or not row:
                continue
            def _c(idx):
                return '' if len(row) <= idx or row[idx] is None else str(row[idx]).strip()
            wings = _c(0)
            if not wings:
                continue
            mm[wings] = {'baumuster': _c(1), 'now': _c(2), 'file': _c(3)}
        out['manual_map'] = mm

    wb.close()
    return out


def load_rules(path: Path | None = None, xlsx_path: Path | None = None) -> dict:
    """Load rules, merging over DEFAULTS so any missing key falls back.

    Priority (highest wins): model_mapping.xlsx > rules.json > built-in DEFAULTS.
    The xlsx is the primary human-editable source; rules.json remains a fallback.
    """
    p = path or RULES_PATH
    merged = {k: (v.copy() if hasattr(v, 'copy') else v) for k, v in DEFAULTS.items()}
    try:
        data = json.loads(Path(p).read_text(encoding='utf-8'))
        for k in DEFAULTS:
            if k in data and data[k] is not None:
                merged[k] = data[k]
    except Exception:
        pass
    # xlsx overrides json/DEFAULTS for whichever keys it defines
    for k, v in load_rules_from_xlsx(xlsx_path).items():
        if v:  # non-empty sheet only, so an accidentally blank sheet can't wipe a rule
            merged[k] = v
    return merged


def apply_map(s: str, mapping: dict) -> str:
    """Replace occurrences of any mapping key in s (longest key first)."""
    if not mapping:
        return s
    keys = sorted(mapping.keys(), key=len, reverse=True)
    pat = '|'.join(re.escape(k) for k in keys)
    return re.sub(pat, lambda m: mapping[m.group()], s)


RULES = load_rules()
