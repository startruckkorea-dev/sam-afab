"""Load the mandatory-code list from code/mandatory-codes.xlsx (source of truth).

The workbook is hand-edited, so it — not option_codes.py — drives which codes are
mandatory. Layout of the 'Mandatory' sheet (header row skipped):
    A = 카테고리(Category)   all / tractor / rigid / tipper   (informational)
    B = Group                one-of group name; blank = individually mandatory
    C = 코드(Code)           e.g. D2Y
    D = 코드명(Description)
    E = 비고(Note)

"Group" semantics: codes sharing a group name satisfy the requirement if AT LEAST
ONE of them is present (e.g. AEBS = S1P/S1W/S2D/S2N — any one counts). A blank group
means the code is individually required.

Path is overridable via config.json 'mandatory_file' / env MANDATORY_FILE. If the
workbook is missing/broken, falls back to option_codes.MANDATORY_CODES/GROUPS so the
build never breaks.
"""
from __future__ import annotations

import json
import os
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent


def _resolve_path(path=None) -> Path:
    if path:
        p = Path(path)
        return p if p.is_absolute() else (ROOT / p)
    rel = os.environ.get('MANDATORY_FILE') or 'code/mandatory-codes.xlsx'
    try:
        cfg = json.loads((ROOT / 'config.json').read_text(encoding='utf-8'))
        rel = os.environ.get('MANDATORY_FILE') or cfg.get('mandatory_file') or rel
    except Exception:
        pass
    p = Path(rel)
    return p if p.is_absolute() else (ROOT / p)


def _rows_from_xlsx(path: Path):
    """Return list of {cat,group,code,desc,note} from the workbook, or None on failure."""
    try:
        from openpyxl import load_workbook  # lazy
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb['Mandatory'] if 'Mandatory' in wb.sheetnames else wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row:  # header
                continue
            def _c(idx):
                return '' if len(row) <= idx or row[idx] is None else str(row[idx]).strip()
            code = _c(2)
            if not code:
                continue
            rows.append({'cat': _c(0), 'group': _c(1), 'code': code,
                         'desc': _c(3), 'note': _c(4)})
        wb.close()
        return rows or None
    except Exception:
        return None


def _rows_from_option_codes():
    """Fallback rows synthesized from option_codes.MANDATORY_CODES/GROUPS."""
    from option_codes import MANDATORY_CODES, MANDATORY_GROUPS
    code2group = {}
    for g, members in MANDATORY_GROUPS.items():
        for c in members:
            code2group[c] = g
    rows = []
    for code, val in MANDATORY_CODES.items():
        desc, note, cat = (val if isinstance(val, tuple) else (val, '', 'all'))
        rows.append({'cat': cat, 'group': code2group.get(code, ''),
                     'code': code, 'desc': desc, 'note': note})
    return rows


def load_mandatory(path=None) -> dict:
    """Return {rows, desc, set, groups} for the mandatory-code list.

    rows   — raw editable rows (for regenerating the workbook)
    desc   — {code: description}          (for codes.json / the front-end)
    set    — {code, ...} all mandatory codes
    groups — {group_name: {codes}}        one-of groups
    """
    p = _resolve_path(path)
    rows = _rows_from_xlsx(p)
    source = 'xlsx'
    if not rows:
        rows = _rows_from_option_codes()
        source = 'option_codes(fallback)'

    desc, mset, groups = {}, set(), {}
    for r in rows:
        c = r['code']
        mset.add(c)
        if c not in desc or (not desc[c] and r['desc']):
            desc[c] = r['desc']
        if r['group']:
            groups.setdefault(r['group'], set()).add(c)
    return {'rows': rows, 'desc': desc, 'set': mset, 'groups': groups, 'source': source}


if __name__ == '__main__':
    m = load_mandatory()
    print(f"source={m['source']}  codes={len(m['set'])}  groups={len(m['groups'])}")
    for g, members in sorted(m['groups'].items()):
        print(f'  [{g}] one-of: {sorted(members)}')
