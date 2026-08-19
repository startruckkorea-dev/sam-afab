"""Generate code/pto-codes.xlsx — the editable list of codes that mark a PTO vehicle.

PTO 판정은 원래 '코드 설명에 PTO 라는 글자가 있나' 였습니다. 설명 문구에 기대는 판정이라
준비(provision)·제어(control) 코드까지 싸잡아 PTO 로 넘겨 버리고, 어떤 코드가 판정을
바꿨는지 사람이 볼 수도 없었습니다. 이제 이 워크북이 유일한 기준입니다 — WINGS 주문과
SAM 견적서 양쪽 모두 '여기 PTO 코드로 적힌 코드를 하나라도 갖고 있나' 로만 판정합니다.

Layout (코드 관리 화면에서 그대로 편집):
    A = 유형(Type)      PTO코드(pto) / 제외(except)
    B = 코드(Code)
    C = 설명(Description)
    D = 비고(Note)

  * PTO코드 — 이 코드가 있으면 그 차/견적서는 PTO 로 본다.
  * 제외    — PTO 와 관련은 있지만 판정에는 쓰지 않는다(준비·제어 코드 등).
              PTO 가 실제로 달리지 않은 차에도 붙기 때문.

Run:  python backend/build_pto_codes_xlsx.py
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
sys.path.insert(0, str(HERE))

from option_codes import OPTION_CODE_MAP  # noqa: E402
from pto_codes import PTO_DEFAULT_CODES, PTO_DEFAULT_EXCEPTS  # noqa: E402

OUT = ROOT / 'code' / 'pto-codes.xlsx'

NAVY = '1F4E79'
PTO_FILL = 'FCE4D6'      # orange-ish — 판정에 쓰는 코드
EXCEPT_FILL = 'E2EFDA'   # green-ish  — 판정에서 빼는 코드

HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
BOLD = Font(bold=True)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center')


def _header(ws, headers):
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER


def build():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'PTO'
    _header(ws, ['유형(Type)', '코드(Code)', '설명(Description)', '비고(Note)'])

    r = 2
    for code, note in PTO_DEFAULT_CODES.items():
        ws.cell(row=r, column=1, value='PTO코드(pto)')
        ws.cell(row=r, column=2, value=code).font = BOLD
        ws.cell(row=r, column=3, value=OPTION_CODE_MAP.get(code, ''))
        ws.cell(row=r, column=4, value=note)
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if col == 1:
                cell.fill = PatternFill('solid', fgColor=PTO_FILL)
        r += 1

    for code, note in PTO_DEFAULT_EXCEPTS.items():
        ws.cell(row=r, column=1, value='제외(except)')
        ws.cell(row=r, column=2, value=code).font = BOLD
        ws.cell(row=r, column=3, value=OPTION_CODE_MAP.get(code, ''))
        ws.cell(row=r, column=4, value=note)
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if col == 1:
                cell.fill = PatternFill('solid', fgColor=EXCEPT_FILL)
        r += 1

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 56
    ws.column_dimensions['D'].width = 46
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:D{r - 1}'

    ws.cell(row=r + 1, column=1,
            value='※ WINGS 주문·SAM 견적서 양쪽 모두 여기 "PTO코드" 를 하나라도 가지면 PTO 로 봅니다. '
                  'PTO 차량과 비PTO 차량은 서로 비교하지 않으므로, 이 목록이 곧 매칭 기준입니다. '
                  '준비/제어 코드처럼 PTO 가 없는 차에도 붙는 코드는 "제외(except)" 로 두세요. '
                  '이 파일이 없으면 내장 기본값으로 진행합니다(빌드 로그에 경고).').font = Font(italic=True, color='777777')

    # --- Sheet 2: 설명에 PTO 가 들어간 모든 코드 (검토용 원본) -------------------
    ref = wb.create_sheet('설명에_PTO포함(Ref)')
    _header(ref, ['코드(Code)', '설명(Description)', '현재 분류'])
    listed = set(PTO_DEFAULT_CODES) | set(PTO_DEFAULT_EXCEPTS)
    hits = sorted(c for c, d in OPTION_CODE_MAP.items()
                  if c and _mentions_pto(d))
    rr = 2
    for c in hits:
        ref.cell(row=rr, column=1, value=c).font = BOLD
        ref.cell(row=rr, column=2, value=OPTION_CODE_MAP.get(c, ''))
        ref.cell(row=rr, column=3,
                 value='PTO코드' if c in PTO_DEFAULT_CODES
                 else ('제외' if c in PTO_DEFAULT_EXCEPTS else '미분류 — 검토 필요'))
        for col in range(1, 4):
            cell = ref.cell(row=rr, column=col)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
        rr += 1
    ref.column_dimensions['A'].width = 12
    ref.column_dimensions['B'].width = 64
    ref.column_dimensions['C'].width = 20
    ref.freeze_panes = 'A2'
    ref.auto_filter.ref = f'A1:C{rr - 1}'
    ref.cell(row=rr + 1, column=1,
             value='※ 확인용 목록입니다(편집해도 판정에 쓰이지 않음). 코드 사전 설명에 PTO 가 '
                   '들어간 코드 전부 — 새 코드가 여기 "미분류" 로 뜨면 왼쪽 PTO 시트에 '
                   '넣을지 결정하세요.').font = Font(italic=True, color='777777')

    wb.save(OUT)
    unlisted = [c for c in hits if c not in listed]
    print(f'Wrote {OUT}  ({len(PTO_DEFAULT_CODES)} PTO codes + '
          f'{len(PTO_DEFAULT_EXCEPTS)} excepts; {len(hits)} codes mention PTO, '
          f'{len(unlisted)} unclassified)')


def _mentions_pto(desc) -> bool:
    """The old heuristic — kept only to build the review list on sheet 2."""
    import re
    d = str(desc or '')
    return ('PTO' in d
            or re.search(r'(?<![A-Za-z])pto(?![A-Za-z])', d, re.IGNORECASE) is not None
            or re.search(r'power[ -]*take[ -]*off', d, re.IGNORECASE) is not None)


if __name__ == '__main__':
    build()
