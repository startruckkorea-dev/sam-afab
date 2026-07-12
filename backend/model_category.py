"""Classify each vehicle as tractor / rigid / tipper from its Baumuster.

The category drives which mandatory codes apply (all + the row's category). The
6-digit Baumuster prefix is the canonical classifier — e.g. 964230 = tipper (4153 K
'Kipper'), 964038 = rigid (4153 L), 963425 = tractor (Actros-L LS). This matches the
category notes the user maintains in the mandatory workbook's Legend.

Source of truth: code/model-category.xlsx (A = BM 6-digit prefix, B = category).
Overridable via config.json 'model_category_file' / env MODEL_CATEGORY_FILE.
Prefixes absent from the workbook fall back to the built-in rule below, so a new
Baumuster is still classified (and can be pinned in the workbook later).
"""
from __future__ import annotations

import json
import os
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent

# Built-in fallback rule (also the pre-fill guess for the workbook).
TRACTOR_PREFIXES = {'963425', '964416', '963403', '964424', '983403'}
TIPPER_PREFIXES = {'964230', '964214', '964231'}
VALID_CATS = {'tractor', 'rigid', 'tipper'}


def classify_prefix(prefix: str) -> str:
    """Rule-based category for a 6-digit BM prefix (fallback when not in the table)."""
    if not prefix:
        return ''
    if prefix in TRACTOR_PREFIXES:
        return 'tractor'
    if prefix in TIPPER_PREFIXES:
        return 'tipper'
    if prefix.startswith('964'):   # other 964xxx = rigid
        return 'rigid'
    return ''


def _resolve_path(path=None) -> Path:
    if path:
        p = Path(path)
        return p if p.is_absolute() else (ROOT / p)
    rel = os.environ.get('MODEL_CATEGORY_FILE') or 'code/model-category.xlsx'
    try:
        cfg = json.loads((ROOT / 'config.json').read_text(encoding='utf-8'))
        rel = os.environ.get('MODEL_CATEGORY_FILE') or cfg.get('model_category_file') or rel
    except Exception:
        pass
    p = Path(rel)
    return p if p.is_absolute() else (ROOT / p)


def load_model_category(path=None) -> dict:
    """Return {bm_prefix(6-digit): category} from code/model-category.xlsx.

    Empty dict if the workbook is missing/broken (callers then rely on
    classify_prefix). Only rows with a valid category are kept.
    """
    p = _resolve_path(path)
    out: dict = {}
    try:
        from openpyxl import load_workbook  # lazy
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or row[0] is None:
                continue
            m = re.match(r'(\d{6})', str(row[0]).strip())
            if not m:
                continue
            cat = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ''
            if cat in VALID_CATS:
                out[m.group(1)] = cat
        wb.close()
    except Exception:
        return {}
    return out


def category_for_baumuster(bm, table: dict) -> str:
    """Category for a full Baumuster (e.g. 96342512): table wins, else the rule."""
    m = re.match(r'(\d{6})', str(bm or '').strip())
    if not m:
        return ''
    prefix = m.group(1)
    return table.get(prefix) or classify_prefix(prefix)


if __name__ == '__main__':
    tbl = load_model_category()
    print(f'loaded {len(tbl)} prefixes from workbook')
    for k, v in sorted(tbl.items()):
        print(f'  {k} -> {v}')
