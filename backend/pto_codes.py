"""PTO 코드 목록 — code/pto-codes.xlsx 를 읽고, 없으면 내장 기본값으로 폴백.

PTO 여부는 '코드 설명에 PTO 라는 글자가 있나' 로 판정하던 것을 그만두고, 이 표 하나로
정한다. WINGS 주문과 SAM 견적서 양쪽에 같은 표를 적용하므로 판정이 대칭이고, 어떤
코드가 판정을 갈랐는지 사람이 워크북에서 그대로 볼 수 있다.

  PTO_DEFAULT_CODES   — 있으면 PTO 로 보는 코드
  PTO_DEFAULT_EXCEPTS — PTO 관련이지만 판정에 쓰지 않는 코드(준비·제어 등)

The JS port is docs/lib/refdata.js loadPtoCodes() — keep the two in step.
"""
from __future__ import annotations

from pathlib import Path

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
XLSX_PATH = ROOT / 'code' / 'pto-codes.xlsx'
SHEET = 'PTO'

# 값은 워크북의 '비고' 열로 들어간다 — 왜 이렇게 분류했는지가 파일에 남게.
PTO_DEFAULT_CODES = {
    'N0A': 'PTO 장착 위치',
    'N0B': 'PTO 장착 위치',
    'N0C': 'PTO 장착 위치',
    'N0D': 'PTO 장착 위치',
    'N0E': 'PTO 장착 위치',
    'N0F': 'PTO 구동 방식',
    'N0G': 'PTO 조작 방식',
    'N0H': 'PTO 조작 방식',
    'N1A': 'PTO 모델',
    'N1B': 'PTO 모델',
    'N1C': 'PTO 모델',
    'N1D': 'PTO 모델',
    'N1E': 'PTO 모델',
    'N1F': 'PTO 모델',
    'N1G': 'PTO 모델',
    'N1H': 'PTO 모델',
    'N1I': 'PTO 모델',
    'N1J': 'PTO 모델',
    'N1K': 'PTO 모델',
    'N1L': 'PTO 모델',
    'N1M': 'PTO 모델',
    'N2E': 'PTO 모델',
    'Z5M': 'PTO 사양',
    'Z5S': 'PTO 사양',
    'Z5U': 'PTO 사양',
}

PTO_DEFAULT_EXCEPTS = {
    'N6P': '준비(provision)만 — PTO 가 실제로 달리지 않은 차에도 붙는다',
    'U2K': '제어 코드(변속기 오일 레벨) — PTO 자체를 뜻하지 않는다',
}


def _norm(v) -> str:
    return '' if v is None else str(v).strip().upper()


def load_pto_codes(path: Path | None = None) -> dict:
    """{'codes': set, 'excepts': set} — 워크북이 없거나 비면 내장 기본값."""
    fallback = {'codes': set(PTO_DEFAULT_CODES), 'excepts': set(PTO_DEFAULT_EXCEPTS),
                'source': 'defaults'}
    p = Path(path) if path else XLSX_PATH
    if not p.exists():
        return fallback
    try:
        from openpyxl import load_workbook
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb[SHEET] if SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
        codes, excepts = set(), set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            kind = str(row[0] or '')
            code = _norm(row[1] if len(row) > 1 else '')
            if not code or code in ('-', '—'):
                continue
            if 'except' in kind.lower() or '제외' in kind:
                excepts.add(code)
            else:
                codes.add(code)
        wb.close()
    except Exception:
        return fallback
    if not codes:
        return fallback
    return {'codes': codes, 'excepts': excepts, 'source': str(p)}


def pto_codes_in(codes, table=None) -> set:
    """주어진 코드 집합에서 PTO 를 뜻하는 코드만."""
    t = table or load_pto_codes()
    listed, excepts = t['codes'], t['excepts']
    return {c for c in (codes or ()) if c in listed and c not in excepts}
