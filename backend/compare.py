"""SAM vs WINGS comparison — extracted from streamlit_app.py, Streamlit-free.

The original used st.session_state for user-customizable code sets; here those
default to values derived from OPTION_CODE_MAP / MANDATORY_CODES.
"""
from __future__ import annotations

import json
import os
import re
from datetime import date
from pathlib import Path

import pandas as pd

from option_codes import OPTION_CODE_MAP
from sam_parser import normalize_model
from rules import RULES, apply_map
from mandatory_codes import load_mandatory
from model_category import load_model_category, category_for_baumuster

# Default Factory-Control (exception) code set: prefixes I/O/Z/U + a few extras.
DEFAULT_EXCEPT_CODES = (
    {c for c in OPTION_CODE_MAP if c and c[0] in {'I', 'O', 'Z', 'U'}}
    | {'DUP0', 'A0B', 'E0D', 'E0Q', 'J7G'}
)

# Mandatory codes come from code/mandatory-codes.xlsx (hand-edited source of truth),
# falling back to option_codes.py. MAND_GROUPS are "one-of" groups: a group is met
# if at least one member is present, so a variant swap within a group is NOT a miss.
# MAND_CATS maps code -> {categories}: a code applies to a row only when 'all' is in
# its categories or the row's vehicle category (tractor/rigid/tipper) matches.
_MAND = load_mandatory()
DEFAULT_MAND_CODES = set(_MAND['set'])
MAND_GROUPS = _MAND['groups']
MAND_CATS = _MAND['cats']
_GROUPED_CODES = set().union(*MAND_GROUPS.values()) if MAND_GROUPS else set()

# Baumuster-prefix -> tractor/rigid/tipper (code/model-category.xlsx + rule fallback).
MODEL_CATEGORY = load_model_category()


def _mand_applies(code: str, row_cat: str) -> bool:
    """True if a mandatory code applies to a vehicle of the given category.

    'all' codes always apply. Category-specific codes apply only when the row's
    category matches. When the row category is unknown, only 'all' codes apply
    (so tipper/tractor-only codes are not force-flagged on an unclassified row).
    """
    cats = MAND_CATS.get(code, {'all'})
    if 'all' in cats:
        return True
    return bool(row_cat) and row_cat in cats

_ROOT = Path(__file__).resolve().parent.parent


def _load_cab_map() -> dict:
    """WINGS cab code -> cab variant found in the SAM filename (e.g. F1J -> G5F).

    Source: code/cab.xlsx (A=code, C=cab variant). Path is overridable via
    config.json 'cab_file' / env CAB_FILE. Used as an extra matching signal:
    a WINGS row carrying F1J should match the SAM file whose title contains G5F.
    Every failure is swallowed so a missing/broken workbook never breaks the build.
    """
    rel = os.environ.get('CAB_FILE') or 'code/cab.xlsx'
    try:
        cfg = json.loads((_ROOT / 'config.json').read_text(encoding='utf-8'))
        rel = os.environ.get('CAB_FILE') or cfg.get('cab_file') or rel
    except Exception:
        pass
    path = Path(rel)
    if not path.is_absolute():
        path = _ROOT / path
    out: dict = {}
    try:
        from openpyxl import load_workbook  # lazy
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or not row[0]:
                continue
            code = str(row[0]).strip().upper()
            variant = str(row[2]).strip().upper() if len(row) > 2 and row[2] else ''
            if code and variant:
                out[code] = variant
        wb.close()
    except Exception:
        return {}
    return out


CAB_MAP = _load_cab_map()


def _split_model(s: str):
    m = re.match(r'^(\d+)([A-Z]*)$', s)
    if m:
        return m.group(1), m.group(2)
    return s, ''


def _candidates(entry, prefer_pto: bool):
    """Return the list of candidate data-dicts for the requested PTO variant.

    A sam_map entry is { is_pto(bool): [data, ...] }. Falls back to the other PTO
    variant when the requested one is absent.
    """
    if not isinstance(entry, dict) or not entry:
        return []
    lst = entry.get(prefer_pto) or entry.get(not prefer_pto) or []
    return lst if isinstance(lst, list) else [lst]


def _match_score(data, wings_bm: str, wings_sub: str, expected_cabs=()) -> int:
    """Higher = better SAM candidate for a WINGS row. Cab > Baumuster > Subcategory.

    expected_cabs: cab variants (e.g. {'G5F'}) derived from the WINGS row's cab
    code via CAB_MAP. A SAM file whose title contains one is a strong match.
    """
    score = 0
    if expected_cabs:
        f = str(data.get('file', '')).upper()
        if any(cv and cv in f for cv in expected_cabs):
            score += 3
    if wings_bm and str(data.get('bm', '')).strip() == wings_bm:
        score += 2
    if wings_sub and str(data.get('sub', '')).strip().lower() == wings_sub:
        score += 1
    return score


def _pick(entry, prefer_pto: bool, wings_bm: str = '', wings_sub: str = '', expected_cabs=()):
    """Pick the single best candidate data-dict from an entry (by Cab/Baumuster/Subcategory)."""
    cands = _candidates(entry, prefer_pto)
    if not cands:
        return None
    return max(cands, key=lambda d: _match_score(d, wings_bm, wings_sub, expected_cabs))


def _cab_ok(data, expected_cabs) -> bool:
    """True if the candidate's SAM filename contains one of the expected cab variants."""
    if not data or not expected_cabs:
        return False
    f = str(data.get('file', '')).upper()
    return any(cv and cv in f for cv in expected_cabs)


def _get_sam_data(entry, prefer_pto: bool, wings_bm: str = '', wings_sub: str = ''):
    """Return (codes_set, filename) from a sam_map entry, best-matching candidate."""
    if isinstance(entry, set):
        return entry, ''
    data = _pick(entry, prefer_pto, wings_bm, wings_sub)
    if data:
        return data['codes'], data['file']
    return set(), ''


def _get_sam_models(entry, prefer_pto: bool, wings_bm: str = '', wings_sub: str = ''):
    """Return (model_baumuster, model_now) from the best-matching candidate."""
    data = _pick(entry, prefer_pto, wings_bm, wings_sub)
    if data:
        return data.get('model_baumuster', ''), data.get('model_now', '')
    return '', ''


def _find_sam_data_by_file(sam_maps_list, prefer_pto: bool, file_sub: str):
    """Find a SAM candidate whose filename contains file_sub (manual file pin).

    Searches all priority maps; prefers the requested PTO variant. Used to force a
    WINGS row onto a specific SAM file when auto-matching picked the wrong one.
    """
    fs = (file_sub or '').strip().lower()
    if not fs:
        return None
    for prefer in (prefer_pto, not prefer_pto):
        for _map in sam_maps_list:
            for entry in _map.values():
                if not isinstance(entry, dict):
                    continue
                for d in (entry.get(prefer) or []):
                    if fs in str(d.get('file', '')).lower():
                        return d
    return None


def compare(df_wings: pd.DataFrame, sam_maps_by_month: dict,
            except_codes: set | None = None, mand_codes: set | None = None,
            allcode_custom: dict | None = None) -> pd.DataFrame:
    """Compare WINGS rows against month-keyed SAM maps. Returns a result DataFrame."""
    _exc_set = except_codes if except_codes is not None else DEFAULT_EXCEPT_CODES
    _mand_set = mand_codes if mand_codes is not None else DEFAULT_MAND_CODES
    _allcode_custom = allcode_custom or {}

    # Manual overrides (model_mapping.xlsx '수동매핑'), keyed by normalized WINGS model.
    _manual_norm = {normalize_model(k): v for k, v in (RULES.get('manual_map') or {}).items()}

    sorted_yyyymm = sorted(sam_maps_by_month.keys())

    def _get_sam_maps_for_prod_date(prod_date_raw) -> list:
        if not sorted_yyyymm:
            return []
        if prod_date_raw:
            try:
                prod_dt = pd.to_datetime(str(prod_date_raw), errors='coerce')
                if not pd.isna(prod_dt):
                    prod_yyyymm = prod_dt.year * 100 + prod_dt.month
                    by_distance = sorted(sorted_yyyymm, key=lambda ym: abs(ym - prod_yyyymm))
                    return [sam_maps_by_month[ym] for ym in by_distance]
            except Exception:
                pass
        return [sam_maps_by_month[ym] for ym in reversed(sorted_yyyymm)]

    def _is_fc(c):
        return c in _exc_set or (c and c[0] in ('I', 'O', 'Z', 'U'))

    rows = []
    for _, r in df_wings.iterrows():
        prod_date_raw = r.get('Requested delivery date', '') if 'Requested delivery date' in r.index else ''
        sam_maps_list = _get_sam_maps_for_prod_date(prod_date_raw)

        com = r['Commission no.']
        model_raw = r.get('Model') or r.get('Baumuster', '')
        baumuster_num = r.get('Baumuster', '') if 'Model' in r else ''
        wings_codes = set(r['WINGS_codes'] or [])
        wings_paint = set(r.get('WINGS_paint') or []) if 'WINGS_paint' in r.index else set()
        wings_tyre = set(r.get('WINGS_tyre') or []) if 'WINGS_tyre' in r.index else set()
        model_norm = normalize_model(model_raw)

        # Extra match fields — from WINGS columns (SAM side comes from the word body).
        wings_bm = str(r.get('Baumuster', '') or '').strip()
        wings_sub = str(r.get('Subcategory (ID)', '') or '').strip().lower()

        # Cab signal: WINGS cab code (e.g. F1J) -> variant in SAM title (e.g. G5F).
        expected_cabs = {CAB_MAP[c] for c in wings_codes if c in CAB_MAP}

        is_pto = any(
            'PTO' in OPTION_CODE_MAP.get(c, '').upper() or
            'PTO' in _allcode_custom.get(c, '').upper()
            for c in wings_codes
        ) or bool(r.get('WINGS_has_pto', False))

        # Find the SAM entry whose model matches, scanning months by production-date
        # proximity. Normally the nearest month with a candidate wins, BUT a cab-correct
        # candidate (WINGS cab code -> variant in the SAM title) beats mere month
        # proximity: a matching GigaSpace file one month away is a better comparison than
        # a StreamSpace file in the exact month. Fall back to nearest-with-any if no
        # month has a cab match.
        def _entry_for_model(_map):
            _cand_entry = _map.get(model_norm, {})
            if _candidates(_cand_entry, is_pto):
                return _cand_entry
            num_norm, suf_norm = _split_model(model_norm)
            for k, v in _map.items():
                k_norm = normalize_model(str(k))
                num_k, suf_k = _split_model(k_norm)
                if k_norm == model_norm or (num_k == num_norm and suf_k == suf_norm):
                    if _candidates(v, is_pto):
                        return v
            return {}

        sam_entry = {}
        sam_map = sam_maps_list[0] if sam_maps_list else {}
        _fallback_entry, _fallback_map = None, None
        for _try_map in sam_maps_list:
            _e = _entry_for_model(_try_map)
            if not _candidates(_e, is_pto):
                continue
            if _fallback_entry is None:          # nearest month with any candidate
                _fallback_entry, _fallback_map = _e, _try_map
            if not expected_cabs:                # no cab signal -> nearest month wins
                sam_entry, sam_map = _e, _try_map
                break
            # cab signal present: take the first (nearest) month that has a cab match
            if _cab_ok(_pick(_e, is_pto, wings_bm, wings_sub, expected_cabs), expected_cabs):
                sam_entry, sam_map = _e, _try_map
                break
        else:
            # no month had a cab match -> fall back to nearest month with any candidate
            if _fallback_entry is not None:
                sam_entry, sam_map = _fallback_entry, _fallback_map

        # Refine PTO: if WINGS has a code unique to the PTO SAM variant.
        if not is_pto and isinstance(sam_entry, dict) and True in sam_entry and False in sam_entry:
            _pto_data = _pick(sam_entry, True, wings_bm, wings_sub)
            _npto_data = _pick(sam_entry, False, wings_bm, wings_sub)
            if _pto_data and _npto_data:
                pto_unique = _pto_data['codes'] - _npto_data['codes']
                if wings_codes & pto_unique:
                    is_pto = True

        # Best auto-matched candidate for this row.
        sam_data = _pick(sam_entry, is_pto, wings_bm, wings_sub, expected_cabs)

        # Manual override ('수동매핑'): file pin replaces the matched SAM file entirely
        # (so codes/comparison use it); baumuster/now text overrides just the display.
        _mo = _manual_norm.get(model_norm)
        if _mo and _mo.get('file'):
            _pinned = _find_sam_data_by_file(sam_maps_list, is_pto, _mo['file'])
            if _pinned is not None:
                sam_data = _pinned

        sam_codes = sam_data['codes'] if sam_data else set()
        sam_file = sam_data['file'] if sam_data else ''
        sam_paint = set(sam_data.get('paint') or []) if sam_data else set()
        sam_tyre = set(sam_data.get('tyre') or []) if sam_data else set()

        # Vehicle category (tractor/rigid/tipper) from Baumuster -> which mandatory
        # codes apply. WINGS Baumuster is the vehicle's own factory code; fall back to
        # the matched SAM file's Baumuster. Category-specific codes (e.g. tipper-only
        # J9J) are then only enforced on vehicles of that category.
        _row_cat = category_for_baumuster(
            wings_bm or (sam_data.get('bm', '') if sam_data else ''), MODEL_CATEGORY)
        _eff_mand = {c for c in _mand_set if _mand_applies(c, _row_cat)}
        _eff_grouped = _GROUPED_CODES & _eff_mand

        only_w = sorted(c for c in (wings_codes - sam_codes)
                        if c and not _is_fc(c) and c not in _eff_mand) if sam_codes else []
        only_s = sorted(c for c in (sam_codes - wings_codes)
                        if c and not _is_fc(c) and c not in _eff_mand)
        except_codes_row = sorted(
            c for c in ((wings_codes - sam_codes) | (sam_codes - wings_codes)) if c and _is_fc(c)
        ) if sam_codes else []
        # Mandatory diff: a code present on exactly one side is a miss — restricted to
        # codes that APPLY to this vehicle's category, EXCEPT for "one-of" groups (e.g.
        # AEBS), which are satisfied as long as each side has some member. A grouped
        # member is flagged only when the group's presence differs between sides.
        _only_one_side = sam_codes ^ wings_codes  # symmetric difference
        _ungrouped_flag = {c for c in _only_one_side
                           if c and c in _eff_mand and c not in _eff_grouped}
        _group_flag = set()
        for _members in MAND_GROUPS.values():
            _em = _members & _eff_mand
            if not _em:
                continue
            _s_has, _w_has = bool(sam_codes & _em), bool(wings_codes & _em)
            if _s_has != _w_has:
                _group_flag |= (_only_one_side & _em)
        mand_codes_row = sorted(_ungrouped_flag | _group_flag)

        # Paint + Tyre (CTT) are compared SEPARATELY from the 3-char option codes and
        # shown as their own sections at the top of the detail chart. A paint/tyre
        # mismatch is only counted when a SAM file matched AND both sides carry that
        # category (so a source that simply omits the section is not flagged).
        _paint_mismatch = bool(sam_data and wings_paint and sam_paint
                               and (wings_paint ^ sam_paint))
        _tyre_mismatch = bool(sam_data and wings_tyre and sam_tyre
                              and (wings_tyre ^ sam_tyre))

        # Vehicle / axle / cab / PTO from SAM filename.
        _vehicle = _axle_type = _cab_code = _pto_flag = ''
        if sam_file:
            _veh_m = re.search(r'\b(Actros-L|Actros|Arocs|Atego|eActros|Econic|Unimog)\b',
                               sam_file, re.IGNORECASE)
            if _veh_m:
                _vehicle = _veh_m.group(1)
            _axle_m = re.search(r'\b(\d+x\d+)\b', sam_file, re.IGNORECASE)
            if _axle_m:
                _axle_type = _axle_m.group(1)
            _cab_m = re.search(r'\b([A-Z]\d[A-Z])\b', sam_file)
            if _cab_m:
                _cab_code = _cab_m.group(1)
            if re.search(r'\bPTO\b', sam_file, re.IGNORECASE):
                _pto_flag = 'PTO'
        # Fallback: some SAM filenames carry no cab token (e.g. tipper "4153 K"
        # files "...Arocs 4153 K 8x4 2026-04..."). Derive the cab variant from the
        # WINGS cab code via cab.xlsx (F1B -> C3M) so the column is not left blank.
        if not _cab_code and expected_cabs:
            _cab_code = sorted(expected_cabs)[0]
        if not _pto_flag and is_pto:
            _pto_flag = 'PTO'
        if not _vehicle:
            _mu = str(model_raw).upper()
            for _veh, _kws in RULES['vehicle_keywords'].items():
                if any(k in _mu for k in _kws):
                    _vehicle = _veh
                    break

        # Status: distinguish "no SAM to compare" from "compared but differs".
        if not sam_file:
            sam_status = 'No SAM'
        elif only_s or only_w or _paint_mismatch or _tyre_mismatch:
            sam_status = 'Mismatch'
        else:
            sam_status = 'Match'

        # SAM designations come straight from the matched file (no rule maps):
        #   Baumuster = body 'Vehicle type' (원본/구형), now = filename model (수정/현행).
        _sam_baumuster = sam_data.get('model_baumuster', '') if sam_data else ''
        _sam_now = sam_data.get('model_now', '') if sam_data else ''
        # Manual text override wins for display (works even with no SAM file).
        if _mo:
            if _mo.get('baumuster'):
                _sam_baumuster = _mo['baumuster']
            if _mo.get('now'):
                _sam_now = _mo['now']
        row_dict = {
            'Commission no.': com,
            'Baumuster': r.get('Baumuster', '') if 'Baumuster' in r.index else baumuster_num,
            # WINGS model shown exactly as read (no display substitution).
            'Model(WINGS)': re.sub(r'DNA$', '',
                str(r.get('Model', model_raw) if 'Model' in r.index else model_raw).strip()),
            'Vehicle': _vehicle,
            'Category': _row_cat,
            'Type': _axle_type,
            'Cab': _cab_code,
            'PTO': _pto_flag,
            'SAM Baumuster': _sam_baumuster,
            'SAM now': _sam_now,
            'Changeability Date': '',
            'Until Dealine': '',
            'Production date': r.get('Requested delivery date', '') if 'Requested delivery date' in r.index else '',
            'Only_in_SAM': ','.join(only_s),
            'Only_in_WINGS': ','.join(only_w) if sam_codes else '',
            'Factory Control Codes': ','.join(except_codes_row),
            'Mandatory Codes': ','.join(mand_codes_row),
            '_all_wings_codes': ','.join(sorted(wings_codes)),
            '_all_sam_codes': ','.join(sorted(sam_codes)),
            # Paint / Tyre (CTT) — compared and displayed on their own, above the
            # general codes, in the detail chart.
            '_paint_wings': ','.join(sorted(wings_paint)),
            '_paint_sam': ','.join(sorted(sam_paint)),
            '_tyre_wings': ','.join(sorted(wings_tyre)),
            '_tyre_sam': ','.join(sorted(sam_tyre)),
            'Compared SAM file name': sam_file,
            'SAM Status': sam_status,
        }

        # Changeability date + days until deadline.
        change_raw = r.get('Vehicle alterable until', '')
        change_display = ''
        days_left = ''
        if change_raw:
            try:
                cdt = pd.to_datetime(change_raw, errors='coerce')
                if not pd.isna(cdt):
                    days_left = (cdt.date() - date.today()).days
                    change_display = cdt.strftime('%Y-%m-%d')
                else:
                    s = str(change_raw).strip()
                    if s:
                        if s.lower() in ('done', 'passed'):
                            change_display = days_left = 'Passed'
                        else:
                            change_display = s
            except Exception:
                s = str(change_raw).strip()
                change_display = 'Passed' if s.lower() in ('done', 'passed') else s
                days_left = 'Passed' if s.lower() in ('done', 'passed') else ''

        row_dict['Changeability Date'] = change_display
        row_dict['Until Dealine'] = days_left

        for col in ['Order status financial', 'Order status logistical',
                    'Additional equipment (enumeration)', 'FIN', 'Subcategory (ID)',
                    'Requested delivery date']:
            if col in r.index:
                row_dict[col] = r[col]

        rows.append(row_dict)
    return pd.DataFrame(rows)
