"""Generate code/code-overview.xlsx — a documentation workbook describing the
codebase (architecture, modules, functions, data flow, config).

This is a human-readable "map" of the SAM × AFAB comparison project so a new
maintainer can understand what each file/function does without reading all source.

Run:  python backend/build_code_overview_xlsx.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
OUT = ROOT / 'code' / 'code-overview.xlsx'

# --- palette ---------------------------------------------------------------
NAVY = '1F4E79'
BLUE = '2E75B6'
GREY = 'F2F2F2'
GREEN = '548235'
ORANGE = 'BF8F00'
PURPLE = '7A5195'

HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
TITLE_FONT = Font(color='FFFFFF', bold=True, size=16)
SUB_FONT = Font(color='FFFFFF', bold=True, size=11)
BOLD = Font(bold=True)
ITALIC = Font(italic=True, color='777777')
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center')


def _header(ws, row, cols, fill=NAVY):
    for i, val in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=val)
        c.fill = PatternFill('solid', fgColor=fill)
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = BORDER


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _rows(ws, start, rows):
    r = start
    for row in rows:
        for i, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.alignment = WRAP_TOP
            c.border = BORDER
        r += 1
    return r


def _section_bar(ws, row, ncols, text, fill=BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill('solid', fgColor=fill)
    c.font = SUB_FONT
    c.alignment = Alignment(horizontal='left', vertical='center')


# ---------------------------------------------------------------------------
def sheet_overview(wb):
    ws = wb.active
    ws.title = '개요(Overview)'
    _widths(ws, [26, 100])
    ws.merge_cells('A1:B1')
    t = ws.cell(row=1, column=1, value='SAM × AFAB 비교 시스템 — 코드 개요')
    t.fill = PatternFill('solid', fgColor=NAVY)
    t.font = TITLE_FONT
    t.alignment = CENTER
    ws.row_dimensions[1].height = 34

    info = [
        ('프로젝트', 'SAM(사내 견적 .docx) ↔ WINGS(주문 시스템 export) 옵션코드 비교'),
        ('아키텍처', '"계산은 GitHub Actions에서, 표시는 GitHub Pages에서". 브라우저는 완성된 JSON만 읽어 표시.'),
        ('백엔드(계산)', 'backend/ — 순수 Python. SAM/WINGS 파싱 → 비교 → docs/data.json + codes.json 생성'),
        ('프런트(표시)', 'docs/ — 정적 HTML/CSS/JS. data.json/codes.json을 읽어 테이블 렌더링 (GitHub Pages 루트)'),
        ('자동화', '.github/workflows/build.yml — 매일 06:00 KST 실행 (스크래핑 → 빌드 → 커밋)'),
        ('보안', 'WINGS 자격증명(ID/PW/TOTP)은 GitHub Secrets에만 존재. 클라이언트 코드에 노출 없음.'),
        ('설정', 'config.json — SAM/WINGS 폴더 경로, 코드사전 파일/시트. 우선순위: CLI > 환경변수 > config.json > 기본값'),
        ('모델규칙', 'model_rules/model_mapping.xlsx — WINGS↔SAM 모델 인식표 + 수동 매핑/별칭 (사람이 편집)'),
        ('코드사전', 'code/mbtruck-spec-data.xlsx (code_dict 시트) — 옵션코드→영문설명. 없으면 option_codes.py로 폴백'),
    ]
    r = 3
    for k, v in info:
        kc = ws.cell(row=r, column=1, value=k)
        kc.font = BOLD
        kc.fill = PatternFill('solid', fgColor=GREY)
        kc.alignment = WRAP_TOP
        kc.border = BORDER
        vc = ws.cell(row=r, column=2, value=v)
        vc.alignment = WRAP_TOP
        vc.border = BORDER
        r += 1

    r += 1
    _section_bar(ws, r, 2, '데이터 흐름 (Data Flow)', fill=GREEN)
    r += 1
    flow = [
        '1) sam_files/YYYY_MM/*.docx  ─┐',
        '2) WINGS 스크래핑/업로드      ─┴─►  backend/build_data.py',
        '3) build_data → parse_wings() + load_sam_by_month() → compare()',
        '4) → docs/data.json (비교 결과 rows) + docs/codes.json (코드 설명 사전)',
        '5) docs/index.html + app.js 가 JSON을 읽어 브라우저에 테이블 표시',
        '6) model_rules/model_mapping.xlsx 자동 갱신 (인식 검증용)',
    ]
    for line in flow:
        c = ws.cell(row=r, column=1, value=line)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c.alignment = WRAP_TOP
        c.font = Font(name='Consolas')
        r += 1
    ws.freeze_panes = 'A3'


def sheet_files(wb):
    ws = wb.create_sheet('파일구조(Files)')
    _widths(ws, [16, 40, 10, 12, 64])
    _header(ws, 1, ['영역', '경로', '언어', '라인수', '역할'])
    rows = [
        ('backend', 'backend/build_data.py', 'Python', 257, '오케스트레이터. SAM/WINGS 파싱→compare()→data.json/codes.json 생성. WINGS 최신파일 자동선택, 스크래핑 트리거, 코드사전 로드.'),
        ('backend', 'backend/wings_parser.py', 'Python', 107, 'WINGS CSV/Excel 파싱 → 정규화 DataFrame(Commission별 1행, WINGS_codes 집합, PTO 플래그).'),
        ('backend', 'backend/sam_parser.py', 'Python', 202, 'SAM .docx 파싱 → {정규화모델: {PTO여부: [{codes,file,model_now,model_baumuster,bm,sub}]}}. 모델명 정규화 포함.'),
        ('backend', 'backend/compare.py', 'Python', 308, 'SAM↔WINGS 비교 핵심 로직. 모델/PTO/Baumuster 매칭, Only_in_SAM/WINGS, 공장통제/필수 코드 분류, 상태(Match/Mismatch/No SAM) 산출.'),
        ('backend', 'backend/rules.py', 'Python', 142, '모델 변환 규칙 로더. 우선순위 model_mapping.xlsx > docs/rules.json > 내장 DEFAULTS. 수동매핑/별칭/차종 키워드 제공.'),
        ('backend', 'backend/option_codes.py', 'Python', 2250, '참조 데이터. OPTION_CODE_MAP(≈2,149개 옵션코드→설명), MANDATORY_CODES(21개 필수), MANDATORY_GROUPS. 순수 데이터.'),
        ('backend', 'backend/wings_scraper.py', 'Python', 1774, 'WINGS 자동 다운로드(Playwright+로컬 Chrome 프로필). 로그인/TOTP/Extended Search로 Excel 다운로드.'),
        ('backend', 'backend/build_model_rules_xlsx.py', 'Python', 117, 'model_rules/model_mapping.xlsx 생성. 인식모델 대조표(자동) + 수동매핑/매칭별칭(편집) 시트.'),
        ('docs', 'docs/index.html', 'HTML', 94, 'GitHub Pages 진입점. 테이블/필터 UI 골격.'),
        ('docs', 'docs/app.js', 'JS', 644, 'data.json/codes.json 로드 → 테이블 렌더링, 필터/검색/코드 설명 팝업.'),
        ('docs', 'docs/style.css', 'CSS', '-', '프런트 스타일.'),
        ('docs', 'docs/data.json', 'JSON', '(생성)', 'Actions가 생성·커밋하는 비교 결과 (generated_at, columns, summary, rows).'),
        ('docs', 'docs/codes.json', 'JSON', '(생성)', 'Actions가 생성·커밋하는 코드 설명 사전 (options, mandatory).'),
        ('docs', 'docs/rules.json', 'JSON', '-', '모델 변환 규칙 폴백 파일 (xlsx가 우선).'),
        ('config', 'config.json', 'JSON', '-', 'SAM/WINGS 폴더 경로, 코드사전 파일/시트 설정.'),
        ('설정', 'model_rules/model_mapping.xlsx', 'XLSX', '-', 'WINGS↔SAM 모델 인식표 + 수동매핑/별칭 (사람 편집·자동 갱신).'),
        ('참조', 'code/mbtruck-spec-data.xlsx', 'XLSX', '-', '옵션코드 스펙/사전 원본 (code_dict 시트 = 코드→영문설명).'),
        ('CI', '.github/workflows/build.yml', 'YAML', '-', '매일 06:00 KST cron + 수동 실행. 스크래핑→빌드→data.json/codes.json 커밋.'),
    ]
    _rows(ws, 2, rows)
    ws.freeze_panes = 'A2'


def sheet_functions(wb):
    ws = wb.create_sheet('함수·로직(Functions)')
    _widths(ws, [26, 34, 74])
    _header(ws, 1, ['모듈', '함수 / 심볼', '설명'])
    r = 2

    def block(title, rows):
        nonlocal r
        _section_bar(ws, r, 3, title, fill=BLUE)
        r += 1
        r = _rows(ws, r, [('', a, b) for a, b in rows])

    block('build_data.py — 오케스트레이터', [
        ('main()', 'CLI 파싱(--wings/--sam-dir/--wings-dir/--scrape/--out). 스크래핑 or WINGS파일 선택 → build() → data.json/codes.json 쓰기 → model_mapping.xlsx 갱신.'),
        ('build(wings_path, sam_root)', 'parse_wings + load_sam_by_month → compare() → DISPLAY_COLS만 추출, summary(total/matched/mismatched/no_sam) 계산.'),
        ('_build_code_dict()', '스펙 워크북 code_dict 시트(B=코드, C=영문설명) 로드. 실패 시 OPTION_CODE_MAP 폴백.'),
        ('_find_latest_wings() / _wings_recency()', 'wings_data/ 여러 파일 중 최신 자동선택. 파일명 내 epoch-ms/날짜 우선, 없으면 mtime.'),
        ('_resolve_dir()', '디렉터리 우선순위 결정: CLI > 환경변수 > config.json > 기본값. 상대경로는 repo 루트 기준.'),
        ('_clean(v)', 'JSON 안전 스칼라 변환(NaN→"", 날짜→YYYY-MM-DD).'),
        ('DISPLAY_COLS', '프런트에 노출할 컬럼 순서 정의.'),
    ])
    block('wings_parser.py', [
        ('parse_wings(file)', 'CSV(우선)/Excel 읽기 → 모델컬럼/옵션컬럼 자동 탐지 → WINGS_codes(집합)·WINGS_has_pto 생성. Commission no. 필수.'),
        ('_extract_codes(text)', '텍스트에서 3~4자 영숫자 코드 정규식 추출(집합).'),
    ])
    block('sam_parser.py', [
        ('parse_single_sam_file()', '.docx zip→document.xml 파싱. Equipment overview/Standard equipment 표에서 코드 추출. 본문 Vehicle type/Baumuster/Subcategory, 파일명 모델 추출. PTO 판정. mapping에 두 번호(now/baumuster) 키로 등록.'),
        ('normalize_model(model)', '모델 문자열 정규화: 축(8x4)/DNA/비영숫자 제거·대문자화. 세대변환은 하지 않음(실번호 매칭).'),
        ('load_sam_from_folder(folder)', '폴더 내 SAM 파일 전체 파싱 → mapping. reverse_aliases 별칭 자동 생성.'),
        ('load_sam_by_month(root)', 'root/YYYY_MM/ 하위폴더 스캔 → {yyyymm(int): mapping}.'),
    ])
    block('compare.py — 비교 핵심', [
        ('compare(df_wings, sam_maps_by_month, ...)', 'WINGS 각 행을 생산월 근접 SAM맵과 비교. 모델/PTO/Baumuster·Subcategory로 최적 SAM 후보 선택 → 결과 DataFrame.'),
        ('_get_sam_maps_for_prod_date()', '요청 납기월과 가까운 순으로 SAM 월맵 정렬(없으면 최신순).'),
        ('_pick / _candidates / _match_score', '엔트리에서 PTO변형·Baumuster·Subcategory 점수로 최적 SAM 후보 선택.'),
        ('_find_sam_data_by_file()', '수동매핑 파일 지정 시 파일명 부분일치로 SAM 후보 강제 선택.'),
        ('is_pto / PTO 재판정', 'WINGS 코드/플래그로 PTO 판정 후, PTO변형 고유코드가 WINGS에 있으면 재확정.'),
        ('Only_in_SAM/WINGS, Factory Control, Mandatory', '차집합에서 공장통제(I/O/Z/U 접두 등)·필수코드 분류. 상태 Match/Mismatch/No SAM.'),
        ('Changeability Date / Until Dealine', 'Vehicle alterable until → 표시일자 + 마감까지 남은 일수(또는 Passed).'),
        ('DEFAULT_EXCEPT_CODES / DEFAULT_MAND_CODES', '기본 공장통제/필수 코드 집합(사용자 커스터마이즈 대체).'),
    ])
    block('rules.py — 모델 규칙', [
        ('load_rules()', 'DEFAULTS 위에 rules.json, 그 위에 model_mapping.xlsx 병합(빈 시트는 무시).'),
        ('load_rules_from_xlsx()', 'xlsx의 매칭별칭/옵션/수동매핑 시트 파싱(openpyxl 지연 로드, 실패 무시).'),
        ('apply_map(s, mapping)', '가장 긴 키 우선으로 문자열 치환.'),
        ('RULES / DEFAULTS', '전역 규칙 딕셔너리 및 내장 기본값(reverse_aliases, vehicle_keywords, manual_map 등).'),
    ])
    block('build_model_rules_xlsx.py', [
        ('build()', 'docs/data.json 기반 인식모델_대조표(자동) + 수동매핑 + 매칭별칭 시트로 model_mapping.xlsx 생성.'),
        ('_ref_rows()', 'data.json에서 (차종,WINGS,SAM번호,상태 등) 중복 제거 행 생성.'),
    ])
    block('wings_scraper.py', [
        ('download_wings_excel(months, ...)', 'Playwright+전용 Chrome 프로필로 WINGS 로그인(TOTP)·Extended Search·Excel 다운로드. Actions/로컬 스크래핑용.'),
    ])
    ws.freeze_panes = 'A2'


def sheet_data(wb):
    ws = wb.create_sheet('데이터·컬럼(Data)')
    _widths(ws, [30, 16, 76])
    _header(ws, 1, ['항목', '타입', '설명'])
    r = 2

    def block(title, rows, fill=BLUE):
        nonlocal r
        _section_bar(ws, r, 3, title, fill=fill)
        r += 1
        r = _rows(ws, r, [(a, b, c) for a, b, c in rows])

    block('docs/data.json — 최상위', [
        ('generated_at', 'ISO datetime', '빌드 생성 시각(UTC).'),
        ('wings_file', 'string', '사용한 WINGS 파일명.'),
        ('columns', 'string[]', '실제 존재하는 표시 컬럼 목록.'),
        ('summary', 'object', 'total / matched / mismatched / no_sam / sam_months.'),
        ('rows', 'object[]', '비교 결과 행 배열 (아래 컬럼).'),
    ])
    block('data.json rows[*] — 주요 컬럼', [
        ('Commission no.', 'string', 'WINGS 주문 번호(키).'),
        ('Baumuster / Model(WINGS)', 'string', 'WINGS 공장코드 / 모델명.'),
        ('Vehicle / Type / Cab / PTO', 'string', 'SAM 파일명에서 추출한 차종/축형식/캡코드/PTO.'),
        ('SAM Baumuster / SAM now', 'string', 'SAM 원본(본문 Vehicle type) / 현행(파일명) 번호.'),
        ('Only_in_SAM / Only_in_WINGS', 'csv', '한쪽에만 있는 옵션코드(공장통제·필수 제외).'),
        ('Factory Control Codes', 'csv', 'I/O/Z/U 접두 등 공장통제(예외) 코드.'),
        ('Mandatory Codes', 'csv', '한쪽에만 있는 필수코드.'),
        ('Changeability Date / Until Dealine', 'string/int', '변경가능 마감일 / 남은 일수(또는 Passed).'),
        ('SAM Status', 'enum', 'Match / Mismatch / No SAM.'),
        ('Compared SAM file name', 'string', '비교에 사용된 SAM 파일명.'),
        ('_all_wings_codes / _all_sam_codes', 'csv', '전체 코드 집합(팝업/디버그용).'),
    ])
    block('docs/codes.json', [
        ('options', 'map', '옵션코드 → 영문설명(code_dict 또는 OPTION_CODE_MAP).'),
        ('mandatory', 'map', '필수코드 → 설명.'),
    ], fill=GREEN)
    block('option_codes.py — 참조 데이터', [
        ('OPTION_CODE_MAP', 'dict (~2,149)', '옵션코드 → 영문 설명.'),
        ('MANDATORY_CODES', 'dict (21)', '필수코드 → (설명, 비고, 카테고리 all/tractor/rigid/tipper).'),
        ('MANDATORY_GROUPS', 'dict', '그룹 중 하나만 있으면 되는 코드 집합(예: AEBS).'),
    ], fill=ORANGE)
    ws.freeze_panes = 'A2'


def sheet_config(wb):
    ws = wb.create_sheet('설정·실행(Config)')
    _widths(ws, [30, 90])
    r = 1
    _section_bar(ws, r, 2, 'config.json 키', fill=NAVY)
    r += 1
    _header(ws, r, ['키', '설명'])
    r += 1
    r = _rows(ws, r, [
        ('sam_dir', 'SAM 원본 폴더(기본 sam_files). 절대/상대경로.'),
        ('wings_dir', 'WINGS export 폴더(기본 wings_data). 최신파일 자동선택.'),
        ('code_dict_file', '코드사전 워크북 경로(code/mbtruck-spec-data.xlsx).'),
        ('code_dict_sheet', '코드사전 시트명(code_dict, B=코드 C=영문설명).'),
    ])
    r += 1
    _section_bar(ws, r, 2, '우선순위', fill=BLUE)
    r += 1
    r = _rows(ws, r, [
        ('디렉터리 경로', 'CLI(--sam-dir/--wings) > 환경변수(SAM_DIR/WINGS_DIR) > config.json > 기본값'),
        ('모델 규칙', 'model_mapping.xlsx > docs/rules.json > 내장 DEFAULTS'),
        ('코드사전', 'code_dict 시트 > option_codes.OPTION_CODE_MAP'),
    ])
    r += 1
    _section_bar(ws, r, 2, '로컬 실행 명령', fill=GREEN)
    r += 1
    cmds = [
        ('pip install -r requirements.txt', '의존성 설치'),
        ('python backend/build_data.py --wings path/to/wings.xlsx', '지정한 WINGS 파일로 빌드'),
        ('python backend/build_data.py', 'wings_data/ 최신 파일 자동 사용'),
        ('python backend/build_data.py --scrape 2026_04 2026_05', 'WINGS 스크래핑부터(자격증명 필요)'),
        ('python -m http.server -d docs 8000', '결과를 정적서버로 확인(localhost:8000)'),
        ('python backend/build_model_rules_xlsx.py', '모델 인식표 xlsx 재생성'),
        ('python backend/build_code_overview_xlsx.py', '이 코드 개요 xlsx 재생성'),
    ]
    _header(ws, r, ['명령', '설명'])
    r += 1
    for cmd, desc in cmds:
        c1 = ws.cell(row=r, column=1, value=cmd)
        c1.font = Font(name='Consolas')
        c1.alignment = WRAP_TOP
        c1.border = BORDER
        c2 = ws.cell(row=r, column=2, value=desc)
        c2.alignment = WRAP_TOP
        c2.border = BORDER
        r += 1
    r += 1
    _section_bar(ws, r, 2, 'CI (.github/workflows/build.yml)', fill=PURPLE)
    r += 1
    r = _rows(ws, r, [
        ('스케줄', '매일 06:00 KST (cron UTC 21:00) + 수동 실행(months 입력).'),
        ('단계', 'checkout → Python 3.11 → 의존성 → (스크래핑 시)Playwright → build_data.py → data.json/codes.json 커밋.'),
        ('Secrets', 'WINGS_USER / WINGS_PASSWORD / WINGS_TOTP_SECRET (코드에 노출 금지).'),
    ])


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sheet_overview(wb)
    sheet_files(wb)
    sheet_functions(wb)
    sheet_data(wb)
    sheet_config(wb)
    wb.save(OUT)
    print(f'Wrote {OUT}')


if __name__ == '__main__':
    main()
