"""(Re)format code/mandatory-codes.xlsx — the editable mandatory-code list.

The workbook is the hand-edited SOURCE OF TRUTH (read at build time by
mandatory_codes.load_mandatory). This script reads that same data (xlsx-first,
option_codes.py fallback) and rewrites the sheet with consistent styling, so
running it PRESERVES your edits — it does not revert to the hardcoded list.

Layout of the 'Mandatory' sheet:
    A = 카테고리(Category)   all / tractor / rigid / tipper
    B = Group                one-of group name (blank = individually mandatory)
    C = 코드(Code)
    D = 코드명(Description)
    E = 비고(Note)

Run:  python backend/build_mandatory_codes_xlsx.py
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
sys.path.insert(0, str(HERE))

from mandatory_codes import load_mandatory  # noqa: E402

OUT = ROOT / 'code' / 'mandatory-codes.xlsx'

NAVY = '1F4E79'
CAT_FILL = {
    'all': 'E2EFDA', 'tractor': 'DDEBF7', 'rigid': 'FCE4D6', 'tipper': 'FFF2CC',
}
GROUP_FILL = 'EDE7F6'

HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
BOLD = Font(bold=True)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center')

CAT_ORDER = {'all': 0, 'tractor': 1, 'rigid': 2, 'tipper': 3}


def build():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    data = load_mandatory()
    rows = data['rows']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Mandatory'
    headers = ['카테고리(Category)', 'Group', '코드(Code)', '코드명(Description)', '비고(Note)']
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER

    # Stable order: category, then group (grouped rows together), then code.
    ordered = sorted(rows, key=lambda r: (CAT_ORDER.get(r['cat'], 99), r['group'], r['code']))
    r = 2
    for row in ordered:
        ws.cell(row=r, column=1, value=row['cat'])
        ws.cell(row=r, column=2, value=row['group'])
        ws.cell(row=r, column=3, value=row['code']).font = BOLD
        ws.cell(row=r, column=4, value=row['desc'])
        ws.cell(row=r, column=5, value=row['note'])
        for col in range(1, 6):
            cell = ws.cell(row=r, column=col)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
        if row['cat'] in CAT_FILL:
            ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor=CAT_FILL[row['cat']])
        if row['group']:
            ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=GROUP_FILL)
        r += 1

    for col, w in zip('ABCDE', [18, 16, 12, 56, 44]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:E{r - 1}'

    # --- Legend sheet --------------------------------------------------------
    lg = wb.create_sheet('설명(Legend)')
    lg.column_dimensions['A'].width = 16
    lg.column_dimensions['B'].width = 70
    lg.append(['항목', '설명'])
    for i in (1, 2):
        c = lg.cell(row=1, column=i)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
    for k, v in [
        ('all', '모든 차량 필수 (All vehicle mandatory)'),
        ('tractor', '트랙터 전용 필수 (BM 963425, 964416, 963403, 964424)'),
        ('rigid', '리지드 전용 필수 (BM 964XXX)'),
        ('tipper', '티퍼 전용 필수 (BM 964230, 964214)'),
        ('', ''),
        ('Group', '같은 Group 이름끼리는 그 중 하나만 있으면 충족 (one-of). 빈칸이면 개별 필수.'),
    ]:
        lg.append([k, v])
    for g, members in sorted(data['groups'].items()):
        lg.append([g, ', '.join(sorted(members))])
    for row in lg.iter_rows(min_row=2, max_row=lg.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = WRAP_TOP
            cell.border = BORDER
    lg.cell(row=2, column=1).font = BOLD

    wb.save(OUT)
    print(f"Wrote {OUT}  ({len(rows)} rows, {len(data['groups'])} groups, "
          f"source={data['source']})")


if __name__ == '__main__':
    build()
