"""Generate model_rules/model_mapping.xlsx — the WINGS↔SAM model-recognition table.

Since matching is now data-driven (each SAM word file carries both its numbers —
filename = 'SAM now', body 'Vehicle type' = 'SAM Baumuster'), this workbook is a
VERIFICATION view of what got recognized plus the rule sheets:

  * 인식모델_대조표  — auto-generated from docs/data.json: which WINGS model matched
    which SAM file, in the dashboard's own columns (Model · Type · Axle · Cab · MY ·
    PTO) with both SAM numbers and the match status. (read-only view)
  * 미매칭_진단 — auto-generated: which WINGS orders found no SAM and why.
  * the rule sheets (정규화_과거번호 · 모델별칭 · 이전모델 · 현재모델 · WINGS표시치환 ·
    차종키워드 · 옵션), read back at build time by rules.load_rules_from_xlsx().
    Matching keys come from 정규화_과거번호 · 모델별칭 · 옵션(normalize_28xx_to_26xx);
    이전모델/현재모델 are generation neighbours (different trucks) and stay informational.

The hand-correction sheets (수동매핑 · 매칭_별칭(수동)) were dropped — the SAM documents
carry the numbers and codes matching needs, so there was nothing left to pin by hand.

Run:  python backend/build_model_rules_xlsx.py   (or it is refreshed by build_data.py)
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from rules import load_rules

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
DATA_PATH = ROOT / 'docs' / 'data.json'
OUT_DIR = ROOT / 'model_rules'
OUT_PATH = OUT_DIR / 'model_mapping.xlsx'

DIAG_SHEET_TITLE = '미매칭_진단'
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(color='FFFFFF', bold=True)
ALIAS_FILL = PatternFill('solid', fgColor='7A5195')


def _style_header(ws, ncols, fill=HEADER_FILL):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'


def _autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# Model year lives only in the codes: V8Q..V8Z = model year 0..9 (V8W -> 2026).
# data.json carries no MY column — the dashboard derives it the same way (app.js).
_MY_DIGIT = {f'V8{c}': i for i, c in enumerate('QRSTUVWXYZ')}


def _model_year(r) -> str:
    for key in ('_all_sam_codes', '_all_wings_codes'):
        for code in str(r.get(key, '') or '').split(','):
            code = code.strip()
            if code in _MY_DIGIT:
                return str(2020 + _MY_DIGIT[code])
    return ''


def _ref_rows():
    """Distinct recognized models from docs/data.json — dashboard columns + SAM file."""
    if not DATA_PATH.exists():
        return []
    data = json.loads(DATA_PATH.read_text(encoding='utf-8'))
    seen = {}
    for r in data.get('rows', []):
        wings = str(r.get('Model(WINGS)', '') or '').strip()
        if not wings:
            continue
        row = (
            str(r.get('Vehicle', '') or '').strip(),          # Model
            wings,                                            # Type
            str(r.get('Type', '') or '').strip(),             # Axle
            str(r.get('Cab', '') or '').strip(),
            _model_year(r),
            str(r.get('PTO', '') or '').strip(),
            str(r.get('SAM Baumuster', '') or '').strip(),
            str(r.get('SAM now', '') or '').strip(),
            str(r.get('Baumuster', '') or '').strip(),
            str(r.get('Subcategory (ID)', '') or '').strip(),
            str(r.get('SAM Status', '') or '').strip(),
            Path(str(r.get('Compared SAM file name', '') or '')).name,
        )
        seen.setdefault(row, row)
    return sorted(seen.values(), key=lambda k: (k[0], k[1], k[11]))


DIAG_COLS = ['구분', 'Model(WINGS)', 'Cab', 'PTO', '생산월', '대수',
             'SAM 파일 / 폴더', '원인 · 다음 할 일']


def _diag_rows():
    """Why rows found no SAM — grouped by model/cab/PTO/production month.

    The browser build (docs/app.js) writes the same sheet and can also list the
    SAM files it skipped; here only docs/data.json is available.
    """
    if not DATA_PATH.exists():
        return []
    data = json.loads(DATA_PATH.read_text(encoding='utf-8'))
    have = {int(m) for m in data.get('summary', {}).get('sam_months', [])}
    groups = {}
    for r in data.get('rows', []):
        if str(r.get('SAM Status', '')) != 'No SAM':
            continue
        ym = str(r.get('Production date', '') or '')[:7]
        key = (str(r.get('Model(WINGS)', '') or ''), str(r.get('Cab', '') or ''),
               str(r.get('PTO', '') or ''), ym)
        groups[key] = groups.get(key, 0) + 1
    out = []
    for (model, cab, pto, ym), n in sorted(groups.items(), key=lambda kv: -kv[1]):
        ym_num = int(ym.replace('-', '')) if ym[:4].isdigit() else 0
        note = ('그 생산월 SAM 에 이 모델 번호가 없음 — 파일 추가, 또는 같은 차의 다른 번호면 '
                '모델별칭·정규화_과거번호 에 한 줄 추가') if ym_num in have else (
                '생산월 폴더 자체가 없음 — 01. SAM_files 에 %s 폴더/파일 추가' % (ym or '해당 월'))
        out.append(['WINGS 미매칭', model, cab, pto, ym, n, '', note])
    for ym in sorted(have):
        out.append(['생산월 폴더', '', '', '', '%d-%02d' % (ym // 100, ym % 100), '', '',
                    '읽은 생산월'])
    return out


def build():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rules = load_rules()
    wb = Workbook()

    # --- Sheet 1: recognition table (auto) -----------------------------------
    ws = wb.active
    ws.title = '인식모델_대조표'
    # 앞 6개는 대시보드 목록과 같은 항목·순서(Model · Type · Axle · Cab · MY · PTO).
    cols = ['Model(차종)', 'Type(WINGS 모델)', 'Axle(축)', 'Cab(캡)', 'MY', 'PTO',
            'SAM Baumuster(원본)', 'SAM now(수정)', 'Baumuster', 'Subcategory',
            '매칭상태', 'SAM 파일']
    ws.append(cols)
    for row in _ref_rows():
        ws.append(list(row))
    _style_header(ws, len(cols))
    _autofit(ws, [14, 16, 9, 9, 7, 7, 18, 16, 12, 11, 12, 52])

    # --- Sheets 2..: all model-matching rules (single source of truth) --------
    # Every matching rule now lives in this workbook so nothing is hardcoded in the
    # repo. rules.py reads these sheet names back at build time (xlsx wins).
    def _map_sheet(title, key, headers, widths, note=''):
        s = wb.create_sheet(title)
        s.append(headers)
        for k, v in (rules.get(key) or {}).items():
            s.append([k, ', '.join(v) if isinstance(v, list) else str(v)])
        _style_header(s, len(headers), fill=ALIAS_FILL)
        _autofit(s, widths)
        if note:
            s.cell(row=1, column=len(headers) + 2, value=note).font = Font(italic=True, color='777777')

    _map_sheet('정규화_과거번호', 'normalize_historic',
               ['과거/대체 번호', '정규 번호'], [16, 16],
               '※ 서로 다른 세대 번호를 같은 것으로 정규화 (예: 3253→4153).')
    _map_sheet('이전모델', 'previous_model',
               ['현재 번호', '이전 세대 번호'], [16, 16],
               '※ 세대 이웃 정보(참고용). 서로 다른 차라서 매칭 키로는 쓰지 않는다.')
    _map_sheet('현재모델', 'current_model',
               ['이전 번호', '현재 세대 번호'], [16, 16],
               '※ 세대 이웃 정보(참고용). 서로 다른 차라서 매칭 키로는 쓰지 않는다.')
    _map_sheet('WINGS표시치환', 'wings_display_replace',
               ['WINGS 표기', '표시로 치환'], [18, 18],
               '※ 화면 표기용 치환 (예: "2651 LS"→"2851 LS").')
    _map_sheet('차종키워드', 'vehicle_keywords',
               ['차종(Vehicle)', '매칭 번호들(쉼표 구분)'], [16, 60],
               '※ 번호로 차종(Actros/Arocs 등)을 분류.')
    _map_sheet('모델별칭', 'reverse_aliases',
               ['SAM 번호', '같은 차의 WINGS 번호들(쉼표 구분)'], [16, 40],
               '※ 같은 차를 가리키는 번호들. 매칭이 이 표로 넓어진다 (예: 3253 → 4153).')

    so = wb.create_sheet('옵션')
    so.append(['옵션', '값'])
    so.append(['normalize_28xx_to_26xx',
               'true' if rules.get('normalize_28xx_to_26xx') else 'false'])
    _style_header(so, 2, fill=ALIAS_FILL)
    _autofit(so, [26, 12])
    so.cell(row=1, column=4, value='※ true/false. 28xx 번호를 26xx로 정규화할지.').font = Font(italic=True, color='777777')

    # --- 미매칭 진단(자동) — 규칙을 고칠 사람이 보는 화면에 원인을 같이 둔다 -------
    wd = wb.create_sheet(DIAG_SHEET_TITLE)
    wd.append(DIAG_COLS)
    for row in _diag_rows():
        wd.append(row)
    _style_header(wd, len(DIAG_COLS))
    _autofit(wd, [14, 16, 8, 7, 11, 7, 46, 70])

    wb.save(OUT_PATH)
    print(f'Wrote {OUT_PATH}  (rows: {ws.max_row - 1})')


if __name__ == '__main__':
    build()
